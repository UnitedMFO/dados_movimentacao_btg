import pandas as pd
import os
from openpyxl import load_workbook
from utilidades import formata_cpj, formata_para_real
from openpyxl.styles import PatternFill, NamedStyle



def criar_relatorio_movimentacoes(nome_arquivo_csv, nome_base_arquivo):
    """Lê um CSV, gera um arquivo Excel e remove o arquivo CSV original."""
    # Define o diretório de destino para salvar o arquivo Excel
    diretorio_relatorios = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'relatorios')
    os.makedirs(diretorio_relatorios, exist_ok=True)  # Cria o diretório "relatorios" se não existir

    # Lê o arquivo CSV
    print(nome_arquivo_csv)
    df_movimentacoes = pd.read_csv(nome_arquivo_csv, delimiter=',', encoding='utf-8', header=0)

    # Define as colunas esperadas
    colunas_esperadas = ['marketName', 'subMarketName', 'asset', 'fundCnpj', 'movementDate',
                         'movementHistory', 'launchType', 'grossValue', 'irValue', 'iofValue',
                         'dueDate', 'index', 'fee', 'issuer', 'accountingGroupCode']

    # Filtra o DataFrame para manter apenas as colunas esperadas
    df_movimentacoes = df_movimentacoes[colunas_esperadas]

    # Gera um nome de arquivo Excel, evitando sobrescrever arquivos existentes
    contador = 1
    nome_arquivo_excel = nome_base_arquivo
    caminho_arquivo_excel = os.path.join(diretorio_relatorios, f"{nome_arquivo_excel}.xlsx")

    while os.path.exists(caminho_arquivo_excel):
        nome_arquivo_excel = f"{nome_base_arquivo} ({contador})"
        caminho_arquivo_excel = os.path.join(diretorio_relatorios, f"{nome_arquivo_excel}.xlsx")
        contador += 1

    # Salva o DataFrame como Excel no diretório correto
    df_movimentacoes.to_excel(caminho_arquivo_excel, index=False)

    # Remove o arquivo CSV original
    os.remove(nome_arquivo_csv)
    print(f"Arquivo Excel criado com sucesso em: {caminho_arquivo_excel}")

    return caminho_arquivo_excel

def formatar_relatorio_movimentacoes(nome_arquivo_excel):
    """Aplica formatação a um arquivo Excel, como formatação de CNPJ e ajustes nos valores brutos."""
    tipos_movimentacoes_negativas = ['CRÉDITO', 'JUROS', 'JUROS S/ CAPITAL', 'RECEBIMENTO DIVIDENDOS', 'RI', 'RS',
                                     'VENCIMENTO DE TÍTULO', 'VENDA', 'AMORTIZAÇÃO', 'RENDIMENTO']

    # Lê o arquivo Excel
    df_movimentacoes = pd.read_excel(nome_arquivo_excel)

    # Formata o CNPJ na coluna 'fundCnpj'
    df_movimentacoes['fundCnpj'] = df_movimentacoes['fundCnpj'].apply(formata_cpj)

    # Ajusta CNPJ e valores com base nas condições
    for index, linha in df_movimentacoes.iterrows():
        if pd.isna(linha['fundCnpj']):
            if linha['subMarketName'] == 'CC':
                df_movimentacoes.at[index, 'fundCnpj'] = 'cash'
            elif linha['subMarketName'] in ['RF', 'ACOES']:
                df_movimentacoes.at[index, 'fundCnpj'] = linha['asset']

        # Torna 'grossValue' negativo se o tipo de lançamento for de crédito ou similares
        if any(tipo in str(linha['launchType']).upper() for tipo in tipos_movimentacoes_negativas):
            df_movimentacoes.at[index, 'grossValue'] = -abs(linha['grossValue'])

    # Salva as mudanças de volta no arquivo Excel
    df_movimentacoes.to_excel(nome_arquivo_excel, index=False)
    print(f"Formatação do relatório concluída com sucesso")


def calcular_valor_liquido(nome_arquivo_excel):
    """Calcula o valor líquido subtraindo os valores de IR e IOF dos valores brutos."""
    df_movimentacoes = pd.read_excel(nome_arquivo_excel)

    # Calcula o valor líquido, subtraindo IR e IOF, se existirem
    df_movimentacoes['netValue'] = df_movimentacoes.apply(
        lambda linha: linha['grossValue'] if pd.isna(linha['irValue']) or pd.isna(linha['iofValue'])
        else linha['grossValue'] - linha['irValue'] - linha['iofValue'],
        axis=1
    )

    # Reordenar colunas para que 'netValue' fique logo após 'iofValue'
    colunas = df_movimentacoes.columns.tolist()
    iof_coluna_index = colunas.index('iofValue')
    colunas.insert(iof_coluna_index + 1, colunas.pop(colunas.index('netValue')))
    df_movimentacoes = df_movimentacoes[colunas]  # Reordena o DataFrame

    # Salva o arquivo Excel atualizado
    df_movimentacoes.to_excel(nome_arquivo_excel, index=False)

    # Aplica formatação monetária nas colunas especificadas
    formata_para_real(nome_arquivo_excel, ['H', 'I', 'J', 'K'])
    print(f"Valor líquido calculado e formatado com sucesso")


def destacar_valores_negativos(nome_arquivo_excel):
    """Destaca em amarelo as células que possuem valores negativos na coluna de valores brutos."""
    # Carrega o arquivo Excel usando openpyxl
    workbook = load_workbook(nome_arquivo_excel)
    planilha = workbook.active

    # Define o preenchimento amarelo para valores negativos
    preenchimento_amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Aplica o preenchimento em células da coluna 'grossValue' (coluna H)
    for linha in planilha.iter_rows(min_row=2, max_row=planilha.max_row, min_col=8, max_col=8):  # Coluna H = 8ª coluna
        for celula in linha:
            if celula.value is not None and celula.value < 0:
                celula.fill = preenchimento_amarelo

    # Salva o arquivo Excel com as formatações aplicadas
    workbook.save(nome_arquivo_excel)
    print(f"Valores negativos destacados no arquivo Excel")


def criar_planilha_resumo(nome_arquivo_excel):
    """Cria uma nova planilha 'Resumo' com fundCnpj, launchType, irValue, iofValue e irTotal, com formatação contábil brasileira."""

    # Carrega o arquivo Excel existente
    df_movimentacoes = pd.read_excel(nome_arquivo_excel)

    # Faz um único groupby para calcular todas as somas de uma vez
    df_resumo = df_movimentacoes.groupby(['fundCnpj', 'launchType']).agg(
        grossValue=('grossValue', 'sum'),
        irValue=('irValue', 'sum'),
        iofValue=('iofValue', 'sum')
    ).reset_index()

    # Calcula irTotal em uma única etapa
    df_resumo['irTotal'] = df_resumo['irValue'] + df_resumo['iofValue']

    # Faz o pivot apenas para grossValue
    df_pivot_gross = df_resumo.pivot_table(
        index='fundCnpj',  # Mantém o fundCnpj como índice
        columns='launchType',  # Cada valor de launchType será uma nova coluna
        values='grossValue',  # Os valores que serão colocados nas células
        aggfunc='sum',  # Soma os valores (caso haja múltiplos)
        fill_value=0  # Preenche valores faltantes com 0
    )

    # Agrupa por fundCnpj para somar os valores de IR e IOF
    df_ir_iof = df_resumo.groupby('fundCnpj').agg(
        irValue=('irValue', 'sum'),
        iofValue=('iofValue', 'sum'),
        irTotal=('irTotal', 'sum')
    ).reset_index()

    # Junta os resultados finais (grossValue com IR e IOF)
    df_final = pd.merge(df_pivot_gross, df_ir_iof, on='fundCnpj', how='left')

    # Salva a nova planilha "Resumo" com pandas
    with pd.ExcelWriter(nome_arquivo_excel, engine='openpyxl', mode='a') as writer:
        df_final.to_excel(writer, sheet_name='Resumo', index=False)

    # Recarrega o arquivo Excel para aplicar formatações
    workbook = load_workbook(nome_arquivo_excel)
    sheet = workbook['Resumo']

    # Cria o estilo contábil brasileiro
    contabilidade_style = NamedStyle(name="contabilidade_brasileira", number_format='R$ #,##0.00')

    # Aplica o estilo de contabilidade apenas às colunas numéricas
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = contabilidade_style.number_format  # Aplica o formato sem recriar o estilo

    # Salva o arquivo Excel com a nova planilha e formatação aplicada
    workbook.save(nome_arquivo_excel)

    print(f"Resumo criado com sucesso no arquivo {nome_arquivo_excel}")






